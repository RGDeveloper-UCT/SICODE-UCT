from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import hashlib

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from app import db
from app.models.coordinacion import (
    ActividadCoordinacion,
    AnexoCoordinacion,
    DocumentoEmitido,
    MovimientoDispositivo,
    PagoCoordinacion,
    RegistroCoordinacion,
    RemisionCoordinacion,
    RemisionExpediente,
    ReporteMonitoreo,
)
from app.services.coordinacion_service import determinar_estado, resolver_expediente, separar_sp_remision


class ImportadorCoordinacion:
    def __init__(self, ruta_archivo, usuario_id, archivo_nombre=None):
        self.ruta_archivo = Path(ruta_archivo)
        self.usuario_id = usuario_id
        self.archivo_nombre = archivo_nombre or self.ruta_archivo.name
        self.lote_importacion = hashlib.sha256(self.ruta_archivo.read_bytes()).hexdigest()
        self.wb = load_workbook(self.ruta_archivo, read_only=True, data_only=True)
        self.resumen = {
            "total": 0, "completos": 0, "pendientes": 0, "sin_vincular": 0,
            "ya_importados": 0, "por_hoja": {}, "advertencias": [],
        }

    @staticmethod
    def _texto(valor):
        if valor is None:
            return None
        if isinstance(valor, float) and valor.is_integer():
            return str(int(valor))
        texto = str(valor).strip()
        return texto or None

    @staticmethod
    def _fecha(valor):
        if valor is None or valor == "":
            return None
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, date):
            return valor
        if isinstance(valor, (int, float)):
            try:
                convertido = from_excel(valor)
                return convertido.date() if isinstance(convertido, datetime) else convertido
            except Exception:
                return None
        texto = str(valor).strip()
        for formato in ("%d/%m/%Y", "%d/%m%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
            try:
                return datetime.strptime(texto, formato).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _decimal(valor):
        if valor is None or valor == "":
            return None
        try:
            return Decimal(str(valor).replace(",", "").strip())
        except (InvalidOperation, ValueError):
            return None

    def _ya_importado(self, hoja, fila):
        return RegistroCoordinacion.query.filter_by(
            lote_importacion=self.lote_importacion, hoja_origen=hoja, fila_origen=fila
        ).first() is not None

    def _contabilizar(self, hoja, estado, ya_importado=False):
        info = self.resumen["por_hoja"].setdefault(
            hoja, {"total": 0, "completos": 0, "pendientes": 0, "sin_vincular": 0, "ya_importados": 0}
        )
        info["total"] += 1
        self.resumen["total"] += 1
        if ya_importado:
            info["ya_importados"] += 1; self.resumen["ya_importados"] += 1
        elif estado == "Completo":
            info["completos"] += 1; self.resumen["completos"] += 1
        elif estado == "Pendiente de vincular":
            info["sin_vincular"] += 1; self.resumen["sin_vincular"] += 1
        else:
            info["pendientes"] += 1; self.resumen["pendientes"] += 1

    def _nuevo_registro(self, tipo, hoja, fila, no_sp=None, rc=None, providencia=None, fecha=None,
                        observaciones=None, usuario_origen=None, estado=None, campos_clave=None):
        expediente, no_sp_norm = resolver_expediente(no_sp)
        estado_final = determinar_estado(expediente, no_sp_norm, campos_clave=campos_clave, estado_preferido=estado)
        return RegistroCoordinacion(
            tipo=tipo, expediente_id=expediente.id if expediente else None, no_sp_referencia=no_sp_norm,
            rc=self._texto(rc), providencia=self._texto(providencia), fecha_recepcion=self._fecha(fecha),
            usuario_id=self.usuario_id, usuario_origen=self._texto(usuario_origen), estado=estado_final,
            observaciones=self._texto(observaciones), origen_registro="IMPORTACION_EXCEL",
            archivo_origen=self.archivo_nombre, lote_importacion=self.lote_importacion,
            hoja_origen=hoja, fila_origen=fila,
        )

    def _iter_filas(self, hoja):
        if hoja not in self.wb.sheetnames:
            self.resumen["advertencias"].append(f"No se encontró la hoja {hoja}.")
            return []
        return self.wb[hoja].iter_rows(min_row=2, values_only=True)

    def _procesar_instalaciones(self, importar):
        hoja = "INSTALACIONES"
        for fila, row in enumerate(self._iter_filas(hoja), start=2):
            if not any(v is not None for v in row[:5]):
                continue
            sp, rc, providencia, descripcion, fecha, usuario = (list(row) + [None] * 6)[:6]
            if self._ya_importado(hoja, fila): self._contabilizar(hoja, None, True); continue
            registro = self._nuevo_registro("INSTALACION", hoja, fila, sp, rc, providencia, fecha,
                                            usuario_origen=usuario, campos_clave=[sp, rc, providencia, fecha])
            self._contabilizar(hoja, registro.estado)
            if importar:
                db.session.add(registro); db.session.flush()
                db.session.add(MovimientoDispositivo(registro_id=registro.id, movimiento="INSTALACION",
                                                      descripcion=self._texto(descripcion) or "EXPEDIENTE"))

    def _procesar_desinstalaciones(self, importar):
        hoja = "DESINTALACIONES"
        for fila, row in enumerate(self._iter_filas(hoja), start=2):
            if not any(v is not None for v in row[:5]): continue
            sp, rc, providencia, descripcion, fecha, folios, usuario = (list(row) + [None] * 7)[:7]
            if self._ya_importado(hoja, fila): self._contabilizar(hoja, None, True); continue
            registro = self._nuevo_registro("DESINSTALACION", hoja, fila, sp, rc, providencia, fecha,
                                            usuario_origen=usuario, campos_clave=[sp, rc, providencia, fecha])
            self._contabilizar(hoja, registro.estado)
            if importar:
                db.session.add(registro); db.session.flush()
                db.session.add(MovimientoDispositivo(registro_id=registro.id, movimiento="DESINSTALACION",
                                                      descripcion=self._texto(descripcion) or "DESINSTALACION",
                                                      folios=self._texto(folios)))

    def _procesar_pagos(self, importar):
        hoja = "PAGOS"
        for fila, row in enumerate(self._iter_filas(hoja), start=2):
            if not any(v is not None for v in row[1:5]): continue
            _, sp, rc, providencia, fecha, folios, pago_desde, pago_hasta, boleta, total, nota = (list(row) + [None] * 11)[:11]
            if self._ya_importado(hoja, fila): self._contabilizar(hoja, None, True); continue
            desde_fecha, hasta_fecha = self._fecha(pago_desde), self._fecha(pago_hasta)
            periodo_texto = self._texto(pago_desde) if pago_desde and not desde_fecha else None
            if pago_hasta and not hasta_fecha:
                periodo_texto = " - ".join(x for x in [periodo_texto, self._texto(pago_hasta)] if x)
            registro = self._nuevo_registro("PAGO", hoja, fila, sp, rc, providencia, fecha,
                                            observaciones=nota, campos_clave=[sp, providencia, fecha, boleta, total])
            self._contabilizar(hoja, registro.estado)
            if importar:
                db.session.add(registro); db.session.flush()
                db.session.add(PagoCoordinacion(registro_id=registro.id, folios=self._texto(folios),
                                                periodo_desde=desde_fecha, periodo_hasta=hasta_fecha,
                                                periodo_texto=periodo_texto, boleta=self._texto(boleta), total=self._decimal(total)))

    def _procesar_anexos(self, importar):
        hoja = "ANEXOS"
        for fila, row in enumerate(self._iter_filas(hoja), start=2):
            if not any(v is not None for v in row[:5]): continue
            sp, rc, providencia, tipo, fecha, folios, escaneado, numero_anexo = (list(row) + [None] * 8)[:8]
            fecha_escaneado = self._fecha(escaneado)
            if self._ya_importado(hoja, fila): self._contabilizar(hoja, None, True); continue
            registro = self._nuevo_registro("ANEXO", hoja, fila, sp, rc, providencia, fecha,
                                            campos_clave=[sp, rc, providencia, tipo, fecha, escaneado])
            self._contabilizar(hoja, registro.estado)
            if importar:
                db.session.add(registro); db.session.flush()
                db.session.add(AnexoCoordinacion(registro_id=registro.id, tipo_anexo=self._texto(tipo),
                                                 folios=self._texto(folios), escaneado=bool(fecha_escaneado),
                                                 fecha_escaneado=fecha_escaneado, numero_anexo=self._texto(numero_anexo)))

    def _procesar_monitoreo(self, importar):
        hoja = "MONITOREOREPORT"
        for fila, row in enumerate(self._iter_filas(hoja), start=2):
            if not any(v is not None for v in row[:5]): continue
            sp, rc, providencia, tipo_doc, fecha, reporte_no, tipo_evento = (list(row) + [None] * 7)[:7]
            if self._ya_importado(hoja, fila): self._contabilizar(hoja, None, True); continue
            registro = self._nuevo_registro("MONITOREO", hoja, fila, sp, rc, providencia, fecha,
                                            campos_clave=[sp, rc, providencia, fecha, reporte_no, tipo_evento])
            self._contabilizar(hoja, registro.estado)
            if importar:
                db.session.add(registro); db.session.flush()
                db.session.add(ReporteMonitoreo(registro_id=registro.id, tipo_documento=self._texto(tipo_doc) or "PROVIDENCIA",
                                                numero_reporte=self._texto(reporte_no), tipo_evento=self._texto(tipo_evento)))

    def _procesar_emitidos(self, importar):
        hoja = "EMITIDOS FIRMADOS"
        for fila, row in enumerate(self._iter_filas(hoja), start=2):
            if not any(v is not None for v in row[:3]): continue
            numero, rc, descripcion, destino, fecha = (list(row) + [None] * 5)[:5]
            if self._ya_importado(hoja, fila): self._contabilizar(hoja, None, True); continue
            registro = self._nuevo_registro("DOCUMENTO_EMITIDO", hoja, fila, rc=rc, fecha=fecha,
                                            campos_clave=[numero, fecha])
            self._contabilizar(hoja, registro.estado)
            if importar:
                db.session.add(registro); db.session.flush()
                db.session.add(DocumentoEmitido(registro_id=registro.id, numero_documento=self._texto(numero) or f"SIN-NUMERO-{fila}",
                                                descripcion=self._texto(descripcion), destino=self._texto(destino)))

    def _procesar_actividades(self, importar):
        hoja = "ACTIVIDADES"
        for fila, row in enumerate(self._iter_filas(hoja), start=2):
            if not any(v is not None for v in row[1:5]): continue
            _, actividad, tipo, fecha, usuario = (list(row) + [None] * 5)[:5]
            if self._ya_importado(hoja, fila): self._contabilizar(hoja, None, True); continue
            registro = self._nuevo_registro("ACTIVIDAD", hoja, fila, fecha=fecha, usuario_origen=usuario,
                                            campos_clave=[actividad, fecha, usuario])
            self._contabilizar(hoja, registro.estado)
            if importar:
                db.session.add(registro); db.session.flush()
                db.session.add(ActividadCoordinacion(registro_id=registro.id, tipo_actividad=self._texto(tipo),
                                                     descripcion=self._texto(actividad) or "Actividad sin descripción"))

    def _procesar_remisiones(self, importar):
        hoja = "EXPEDIENTESREMITIDOS"
        for fila, row in enumerate(self._iter_filas(hoja), start=2):
            if not any(v is not None for v in row): continue
            sp_celda, folios, anexos, foliacion, remitido = (list(row) + [None] * 5)[:5]
            if self._ya_importado(hoja, fila): self._contabilizar(hoja, None, True); continue
            estado = "Completo" if self._texto(remitido) and "REMIT" in self._texto(remitido).upper() else "Pendiente de remisión"
            registro = self._nuevo_registro("REMISION", hoja, fila, estado=estado)
            self._contabilizar(hoja, registro.estado)
            if importar:
                db.session.add(registro); db.session.flush()
                remision = RemisionCoordinacion(registro_id=registro.id, destino="Archivo/Bodega MINGOB")
                db.session.add(remision); db.session.flush()
                partes = separar_sp_remision(sp_celda)
                if not partes and sp_celda is not None: partes = [self._texto(sp_celda)]
                for no_sp in partes:
                    expediente, no_sp_norm = resolver_expediente(no_sp)
                    db.session.add(RemisionExpediente(
                        remision_id=remision.id, expediente_id=expediente.id if expediente else None,
                        no_sp_referencia=no_sp_norm or self._texto(no_sp) or "SIN-SP", folios=self._texto(folios),
                        anexos=self._texto(anexos), estado_foliacion=self._texto(foliacion),
                    ))

    def procesar(self, importar=False):
        procesadores = [self._procesar_instalaciones, self._procesar_desinstalaciones, self._procesar_pagos,
                        self._procesar_anexos, self._procesar_monitoreo, self._procesar_emitidos,
                        self._procesar_actividades, self._procesar_remisiones]
        try:
            for procesador in procesadores: procesador(importar)
            if importar: db.session.commit()
        except Exception:
            db.session.rollback()
            raise
        finally:
            self.wb.close()
        return self.resumen
