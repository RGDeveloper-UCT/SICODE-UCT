from io import BytesIO
from datetime import datetime

from flask import Blueprint, send_file
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.coordinacion import RegistroCoordinacion
from app.models.soporte_tecnico import ServicioSoporteTecnico
from app.security import admin_required
from app.services.bitacora_service import registrar_bitacora


coordinacion_export_bp = Blueprint(
    "coordinacion_export",
    __name__,
    url_prefix="/coordinacion",
)


_COLOR_CABECERA = "17233C"
_COLOR_CABECERA_TEXTO = "FFFFFF"


def _fecha(valor):
    return valor.strftime("%d/%m/%Y") if valor else ""


def _fecha_hora(valor):
    return valor.strftime("%d/%m/%Y %H:%M:%S") if valor else ""


def _lista(valores):
    return "; ".join(str(valor) for valor in (valores or []))


def _folios(registro):
    if registro.folios_recepcion:
        return registro.folios_recepcion
    if registro.tipo == "PAGO" and registro.pago and registro.pago.folios:
        return registro.pago.folios
    if registro.tipo in {"INSTALACION", "DESINSTALACION"} and registro.movimiento_dispositivo and registro.movimiento_dispositivo.folios:
        return registro.movimiento_dispositivo.folios
    if registro.tipo == "ANEXO" and registro.anexo_coordinacion and registro.anexo_coordinacion.folios:
        return registro.anexo_coordinacion.folios
    return ""


def _usuario(registro):
    if registro.usuario_origen:
        return registro.usuario_origen
    if registro.usuario:
        return registro.usuario.nombre
    return ""


def _periodo_pago(pago):
    if not pago:
        return ""
    if pago.periodo_desde and pago.periodo_hasta:
        return f"{_fecha(pago.periodo_desde)} al {_fecha(pago.periodo_hasta)}"
    if pago.periodo_texto:
        return pago.periodo_texto
    if pago.periodo_desde:
        return _fecha(pago.periodo_desde)
    if pago.periodo_hasta:
        return f"Hasta {_fecha(pago.periodo_hasta)}"
    return ""


def _resumen_especifico(registro):
    if registro.tipo == "PAGO" and registro.pago:
        total = f"Q {registro.pago.total:.2f}" if registro.pago.total is not None else ""
        return f"Período: {_periodo_pago(registro.pago)} | Boleta: {registro.pago.boleta or ''} | Total: {total}"
    if registro.tipo in {"INSTALACION", "DESINSTALACION"} and registro.movimiento_dispositivo:
        return f"{registro.movimiento_dispositivo.movimiento}: {registro.movimiento_dispositivo.descripcion or ''}"
    if registro.tipo == "ANEXO" and registro.anexo_coordinacion:
        anexo = registro.anexo_coordinacion
        return f"Tipo: {anexo.tipo_anexo or ''} | Anexo No.: {anexo.numero_anexo or ''} | Escaneado: {'Sí' if anexo.escaneado else 'Pendiente'}"
    if registro.tipo == "MONITOREO" and registro.reporte_monitoreo:
        reporte = registro.reporte_monitoreo
        return f"Documento: {reporte.tipo_documento or ''} | Reporte: {reporte.numero_reporte or ''} | Evento: {reporte.tipo_evento or ''}"
    if registro.tipo == "DOCUMENTO_EMITIDO" and registro.documento_emitido:
        documento = registro.documento_emitido
        return f"Documento: {documento.numero_documento or ''} | Destino: {documento.destino or ''} | {documento.descripcion or ''}"
    if registro.tipo == "ACTIVIDAD" and registro.soporte_tecnico:
        soporte = registro.soporte_tecnico
        return (
            f"Boleta soporte: {soporte.numero_boleta} | Usuario: {soporte.usuario_solicitante} | "
            f"Área: {soporte.coordinacion_area} | Estado: {soporte.estado_legible} | "
            f"Servicios: {_lista(soporte.tipos_servicio)}"
        )
    if registro.tipo == "ACTIVIDAD" and registro.actividad_coordinacion:
        actividad = registro.actividad_coordinacion
        return f"Tipo: {actividad.tipo_actividad or ''} | Área: {actividad.area_apoyo or ''} | {actividad.descripcion or ''}"
    if registro.tipo == "REMISION" and registro.remision_coordinacion:
        remision = registro.remision_coordinacion
        return f"Destino: {remision.destino or ''} | Control: {remision.numero_control or ''} | Expedientes: {len(remision.expedientes_remitidos)}"
    return ""


def _campos_comunes(registro):
    return [
        registro.id,
        _fecha(registro.fecha_recepcion),
        registro.tipo,
        registro.expediente_id or "",
        registro.no_sp_referencia or "",
        registro.rc or "",
        registro.providencia or "",
        registro.persona_entrega or "",
        _folios(registro),
        registro.usuario_id,
        _usuario(registro),
        registro.estado or "",
        registro.observaciones or "",
        registro.origen_registro or "",
        registro.archivo_origen or "",
        registro.lote_importacion or "",
        registro.hoja_origen or "",
        registro.fila_origen or "",
        _fecha_hora(registro.creado_en),
        _fecha_hora(registro.actualizado_en),
    ]


def _configurar_hoja(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    relleno = PatternFill("solid", fgColor=_COLOR_CABECERA)
    fuente = Font(color=_COLOR_CABECERA_TEXTO, bold=True)
    for celda in ws[1]:
        celda.fill = relleno
        celda.font = fuente
        celda.alignment = Alignment(vertical="center")

    for columna in range(1, ws.max_column + 1):
        letra = get_column_letter(columna)
        ancho = 10
        for celda in ws[letra]:
            valor = "" if celda.value is None else str(celda.value)
            ancho = max(ancho, min(len(valor) + 2, 45))
        ws.column_dimensions[letra].width = ancho

    for fila in ws.iter_rows(min_row=2):
        for celda in fila:
            celda.alignment = Alignment(vertical="top", wrap_text=True)


def _agregar_hoja(wb, titulo, encabezados, filas):
    ws = wb.create_sheet(title=titulo)
    ws.append(encabezados)
    for fila in filas:
        ws.append(fila)
    _configurar_hoja(ws)
    return ws


@coordinacion_export_bp.route("/exportar")
@login_required
@admin_required
def exportar_todos():
    registros = RegistroCoordinacion.query.order_by(
        RegistroCoordinacion.fecha_recepcion.desc().nullslast(),
        RegistroCoordinacion.creado_en.desc(),
        RegistroCoordinacion.id.desc(),
    ).all()

    wb = Workbook()
    wb.remove(wb.active)

    comunes = [
        "ID registro", "Fecha recepción", "Tipo", "Expediente ID", "SP", "RC", "Providencia",
        "Quién entrega/remite", "Folios", "Usuario ID", "Registrado por", "Estado", "Observaciones",
        "Origen", "Archivo origen", "Lote importación", "Hoja origen", "Fila origen", "Creado", "Actualizado",
    ]

    _agregar_hoja(
        wb,
        "Todos",
        comunes + ["Resumen específico"],
        [_campos_comunes(registro) + [_resumen_especifico(registro)] for registro in registros],
    )

    pagos = [r for r in registros if r.tipo == "PAGO" and r.pago]
    _agregar_hoja(
        wb,
        "Pagos",
        comunes + ["Período desde", "Período hasta", "Período texto", "Boleta", "Total"],
        [
            _campos_comunes(r) + [
                _fecha(r.pago.periodo_desde),
                _fecha(r.pago.periodo_hasta),
                r.pago.periodo_texto or "",
                r.pago.boleta or "",
                float(r.pago.total) if r.pago.total is not None else "",
            ]
            for r in pagos
        ],
    )

    instalaciones = [r for r in registros if r.tipo == "INSTALACION" and r.movimiento_dispositivo]
    _agregar_hoja(
        wb,
        "Instalaciones",
        comunes + ["Movimiento", "Descripción", "Folios movimiento"],
        [
            _campos_comunes(r) + [
                r.movimiento_dispositivo.movimiento or "",
                r.movimiento_dispositivo.descripcion or "",
                r.movimiento_dispositivo.folios or "",
            ]
            for r in instalaciones
        ],
    )

    desinstalaciones = [r for r in registros if r.tipo == "DESINSTALACION" and r.movimiento_dispositivo]
    _agregar_hoja(
        wb,
        "Desinstalaciones",
        comunes + ["Movimiento", "Descripción", "Folios movimiento"],
        [
            _campos_comunes(r) + [
                r.movimiento_dispositivo.movimiento or "",
                r.movimiento_dispositivo.descripcion or "",
                r.movimiento_dispositivo.folios or "",
            ]
            for r in desinstalaciones
        ],
    )

    anexos = [r for r in registros if r.tipo == "ANEXO" and r.anexo_coordinacion]
    _agregar_hoja(
        wb,
        "Anexos",
        comunes + ["Documento expediente ID", "Tipo anexo", "Anexo No.", "Escaneado", "Fecha escaneado", "Folios anexo"],
        [
            _campos_comunes(r) + [
                r.anexo_coordinacion.documento_expediente_id or "",
                r.anexo_coordinacion.tipo_anexo or "",
                r.anexo_coordinacion.numero_anexo or "",
                "Sí" if r.anexo_coordinacion.escaneado else "No",
                _fecha(r.anexo_coordinacion.fecha_escaneado),
                r.anexo_coordinacion.folios or "",
            ]
            for r in anexos
        ],
    )

    monitoreos = [r for r in registros if r.tipo == "MONITOREO" and r.reporte_monitoreo]
    _agregar_hoja(
        wb,
        "Monitoreo",
        comunes + ["Tipo documento", "Reporte No.", "Evento"],
        [
            _campos_comunes(r) + [
                r.reporte_monitoreo.tipo_documento or "",
                r.reporte_monitoreo.numero_reporte or "",
                r.reporte_monitoreo.tipo_evento or "",
            ]
            for r in monitoreos
        ],
    )

    documentos = [r for r in registros if r.tipo == "DOCUMENTO_EMITIDO" and r.documento_emitido]
    _agregar_hoja(
        wb,
        "Documentos emitidos",
        comunes + ["No. documento", "Destino", "Descripción documento"],
        [
            _campos_comunes(r) + [
                r.documento_emitido.numero_documento or "",
                r.documento_emitido.destino or "",
                r.documento_emitido.descripcion or "",
            ]
            for r in documentos
        ],
    )

    actividades = [r for r in registros if r.tipo == "ACTIVIDAD" and r.actividad_coordinacion]
    _agregar_hoja(
        wb,
        "Actividades",
        comunes + ["Tipo actividad", "Área apoyada", "Descripción actividad"],
        [
            _campos_comunes(r) + [
                r.actividad_coordinacion.tipo_actividad or "",
                r.actividad_coordinacion.area_apoyo or "",
                r.actividad_coordinacion.descripcion or "",
            ]
            for r in actividades
        ],
    )

    soportes = ServicioSoporteTecnico.query.order_by(
        ServicioSoporteTecnico.fecha_hora_solicitud.desc(),
        ServicioSoporteTecnico.id.desc(),
    ).all()
    _agregar_hoja(
        wb,
        "Soporte técnico",
        [
            "ID boleta", "No. boleta", "ID registro coordinación", "Fecha/hora solicitud",
            "Usuario solicitante", "Puesto/cargo", "Coordinación/área", "Técnico asignado",
            "Tipos de servicio", "Gestión usuario", "Hardware", "Software", "Instalación",
            "Traslado", "Revisión", "Otro servicio TI", "Otro instalación", "Otro traslado",
            "Otro revisión", "Tipo equipo", "Otro tipo equipo", "Marca/modelo", "No. serie",
            "Inventario", "IP/nombre equipo", "Solicitud/falla", "Diagnóstico/trabajo",
            "Estado final", "Seguimiento", "Fecha/hora cierre", "Tiempo empleado",
            "Observaciones cierre", "Nombre firma usuario", "Fecha firma usuario",
            "Nombre firma técnico", "Fecha firma técnico", "Creado", "Actualizado",
        ],
        [
            [
                s.id,
                s.numero_boleta,
                s.registro_id,
                _fecha_hora(s.fecha_hora_solicitud),
                s.usuario_solicitante,
                s.puesto_cargo or "",
                s.coordinacion_area,
                s.tecnico_asignado,
                _lista(s.tipos_servicio),
                _lista(s.gestion_usuario_detalles),
                _lista(s.hardware_detalles),
                _lista(s.software_detalles),
                _lista(s.instalacion_detalles),
                _lista(s.traslado_detalles),
                _lista(s.revision_detalles),
                s.otro_servicio_ti or "",
                s.otro_instalacion or "",
                s.otro_traslado or "",
                s.otro_revision or "",
                s.tipo_equipo or "",
                s.tipo_equipo_otro or "",
                s.marca_modelo or "",
                s.numero_serie or "",
                s.inventario or "",
                s.ip_nombre_equipo or "",
                s.descripcion_solicitud or "",
                s.diagnostico_trabajo or "",
                s.estado_legible,
                "Sí" if s.seguimiento else "No",
                _fecha_hora(s.fecha_hora_cierre),
                s.tiempo_empleado or "",
                s.observaciones_cierre or "",
                s.nombre_firma_usuario or "",
                _fecha(s.fecha_firma_usuario),
                s.nombre_firma_tecnico or "",
                _fecha(s.fecha_firma_tecnico),
                _fecha_hora(s.creado_en),
                _fecha_hora(s.actualizado_en),
            ]
            for s in soportes
        ],
    )

    remisiones = [r for r in registros if r.tipo == "REMISION" and r.remision_coordinacion]
    _agregar_hoja(
        wb,
        "Remisiones",
        comunes + ["Destino", "No. control", "Cantidad expedientes"],
        [
            _campos_comunes(r) + [
                r.remision_coordinacion.destino or "",
                r.remision_coordinacion.numero_control or "",
                len(r.remision_coordinacion.expedientes_remitidos),
            ]
            for r in remisiones
        ],
    )

    detalles_remision = []
    for registro in remisiones:
        remision = registro.remision_coordinacion
        for detalle in remision.expedientes_remitidos:
            detalles_remision.append([
                registro.id,
                remision.id,
                _fecha(registro.fecha_recepcion),
                remision.destino or "",
                remision.numero_control or "",
                detalle.id,
                detalle.expediente_id or "",
                detalle.no_sp_referencia or "",
                detalle.folios or "",
                detalle.anexos or "",
                detalle.estado_foliacion or "",
                detalle.observaciones or "",
            ])

    _agregar_hoja(
        wb,
        "Expedientes remision",
        [
            "ID registro", "ID remisión", "Fecha", "Destino", "No. control", "ID detalle",
            "Expediente ID", "SP", "Folios", "Anexos", "Estado foliación", "Observaciones",
        ],
        detalles_remision,
    )

    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    registrar_bitacora(
        accion="EXPORTAR_COORDINACION_EXCEL",
        modulo="Coordinación",
        descripcion=f"Se exportaron {len(registros)} registros de Coordinación a Excel.",
        usuario_id=current_user.id,
        entidad="RegistroCoordinacion",
        datos_posteriores={
            "registros_exportados": len(registros),
            "boletas_soporte_exportadas": len(soportes),
            "formato": "XLSX",
        },
    )

    nombre = f"SICODE_Coordinacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        archivo,
        as_attachment=True,
        download_name=nombre,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
