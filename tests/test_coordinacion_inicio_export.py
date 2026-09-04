from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook
from werkzeug.security import generate_password_hash

from app import create_app, db
from app.models.bitacora import Bitacora
from app.models.coordinacion import PagoCoordinacion, RegistroCoordinacion
from app.models.expediente import Expediente
from app.models.usuario import Usuario


@pytest.fixture()
def app_coordinacion_export():
    app = create_app()
    app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)

    with app.app_context():
        db.drop_all()
        db.create_all()

        admin = Usuario(
            nombre="Admin Coordinación",
            usuario="admin-coord",
            correo="admin-coord@uct.local",
            password_hash=generate_password_hash("Password123", method="pbkdf2:sha256"),
            debe_cambiar_password=False,
            rol="administrador",
            activo=True,
        )
        expediente = Expediente(
            codigo_interno="SICODE-UCT-0200",
            no_sp="200",
            nombre_referencia="Sujeto Exportación",
            estado_administrativo="Activo",
            estado_fisico_documental="Pendiente de verificación",
            expediente_fisico_registrado=True,
            activo=True,
        )
        db.session.add_all([admin, expediente])
        db.session.flush()

        registro = RegistroCoordinacion(
            tipo="PAGO",
            expediente_id=expediente.id,
            no_sp_referencia="200",
            rc="RC-EXPORT",
            providencia="PROV-EXPORT",
            fecha_recepcion=date(2026, 8, 20),
            persona_entrega="Administración",
            folios_recepcion="3",
            usuario_id=admin.id,
            usuario_origen=admin.nombre,
            estado="Completo",
            origen_registro="MANUAL",
        )
        db.session.add(registro)
        db.session.flush()
        db.session.add(PagoCoordinacion(
            registro_id=registro.id,
            folios="3",
            periodo_desde=date(2026, 7, 1),
            periodo_hasta=date(2026, 7, 31),
            boleta="BOLETA-EXPORT",
            total=1500,
        ))
        db.session.commit()

    yield app

    with app.app_context():
        db.session.remove()
        db.drop_all()


@pytest.fixture()
def cliente_admin(app_coordinacion_export):
    cliente = app_coordinacion_export.test_client()
    respuesta = cliente.post(
        "/login",
        data={"usuario": "admin-coord", "password": "Password123"},
        follow_redirects=False,
    )
    assert respuesta.status_code == 302
    assert respuesta.headers["Location"].endswith("/dashboard/") or respuesta.headers["Location"].endswith("/dashboard")
    return cliente


def test_inicio_coordinacion_muestra_layout_y_exportacion(cliente_admin):
    respuesta = cliente_admin.get("/coordinacion/")
    texto = respuesta.get_data(as_text=True)

    assert respuesta.status_code == 200
    assert 'body class="vista-coordinacion-inicio"' in texto
    assert ">REGISTROS<" in texto
    assert "Últimos registros recientes" in texto
    assert "Soporte técnico y actividades" in texto
    assert "Exportar todos los datos" in texto
    assert "/coordinacion/exportar" in texto
    assert "BOLETA-EXPORT" not in texto

    # Los módulos principales mantienen identidad visual propia. Monitoreo deja
    # de ser tarjeta principal y se conserva, junto con Análisis de riesgo,
    # dentro del subpanel documental de Anexos.
    assert "coordinacion_tarjetas.css" in texto
    for clase in (
        "registro-animacion--pago",
        "registro-animacion--instalacion",
        "registro-animacion--desinstalacion",
        "registro-animacion--anexo",
        "registro-animacion--documento_emitido",
        "registro-animacion--actividad",
        "registro-animacion--remision",
    ):
        assert clase in texto
    # Evita una coincidencia parcial con la tarjeta válida `--monitoreo_masivo`.
    assert 'class="tarjeta-registro-coordinacion tarjeta-registro-coordinacion--monitoreo"' not in texto
    assert "subpanel-anexos" in texto
    assert "Reporte de monitoreo" in texto
    assert "Análisis de riesgo" in texto
    assert "/coordinacion/registrar/monitoreo" in texto
    assert "/coordinacion/registrar/analisis-riesgo" in texto
    assert "boton-tarjeta-registro-principal" in texto
    assert "tarjeta-registro-etiqueta" in texto


def test_exportacion_coordinacion_genera_excel_completo(app_coordinacion_export, cliente_admin):
    respuesta = cliente_admin.get("/coordinacion/exportar")

    assert respuesta.status_code == 200
    assert respuesta.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "SICODE_Coordinacion_" in respuesta.headers["Content-Disposition"]

    libro = load_workbook(BytesIO(respuesta.data), data_only=True)
    assert libro.sheetnames == [
        "Todos",
        "Pagos",
        "Instalaciones",
        "Desinstalaciones",
        "Anexos",
        "Monitoreo",
        "Análisis de riesgo",
        "Documentos emitidos",
        "Actividades",
        "Soporte técnico",
        "Remisiones",
        "Expedientes remision",
    ]

    pagos = libro["Pagos"]
    encabezados = [celda.value for celda in pagos[1]]
    assert "Boleta" in encabezados
    assert "Total" in encabezados

    analisis = libro["Análisis de riesgo"]
    encabezados_analisis = [celda.value for celda in analisis[1]]
    assert "Correlativo" in encabezados_analisis
    assert "Anexo No." in encabezados_analisis

    soporte = libro["Soporte técnico"]
    encabezados_soporte = [celda.value for celda in soporte[1]]
    assert "No. boleta" in encabezados_soporte
    assert "Diagnóstico/trabajo" in encabezados_soporte
    assert "Estado final" in encabezados_soporte

    valores = [celda.value for celda in pagos[2]]
    assert "RC-EXPORT" in valores
    assert "PROV-EXPORT" in valores
    assert "BOLETA-EXPORT" in valores
    assert 1500 in valores

    with app_coordinacion_export.app_context():
        auditoria = Bitacora.query.filter_by(accion="EXPORTAR_COORDINACION_EXCEL").one()
        assert auditoria.modulo == "Coordinación"
        assert auditoria.datos_posteriores["registros_exportados"] == 1
        assert auditoria.datos_posteriores["boletas_soporte_exportadas"] == 0
